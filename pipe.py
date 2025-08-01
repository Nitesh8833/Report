# -*- coding: utf-8 -*-
"""
CO Billing Matrix Transformer
- Reads source Excel that has a banner + table.
- Extracts 'Effective as of <date>' from the banner.
- Normalizes columns, splits combined fields, expands Medicaid Provider # into multiple columns.
- Writes a clean Excel that starts at the header row (no banner).
"""

from __future__ import annotations
import re
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill


# ==============================
# Helpers: banner + header find
# ==============================

def _norm(s: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s).strip().lower()) if s is not None else ""


def find_header_row(ws: Worksheet, search_terms: Optional[List[str]] = None,
                    scan_rows: int = 50) -> int:
    """Locate the most likely header row (1-based)."""
    if search_terms is None:
        search_terms = ["facility", "npi", "medicare provider", "location name", "billing name"]
    norm_terms = {_norm(t) for t in search_terms}

    best_row, best_hits = None, -1
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        vals = [c.value for c in ws[r]]
        norms = [_norm(v) for v in vals if v is not None]
        hits = sum(any(t in n for t in norm_terms) for n in norms)
        if hits > best_hits:
            best_hits = hits
            best_row = r
    return best_row or 6


def extract_effective_date(ws: Worksheet, search_rows: int = 50, search_cols: int = 40) -> str:
    """
    Robustly find 'Effective as of <date>' (or variants) in the banner area.
    Works with 12/12/2024, 12-12-2024, 12.12.2024, or 'Dec 12, 2024'.
    Also checks cells to the right (merged banners often put the date next door).
    Returns 'MM/DD/YYYY' or ''.
    """
    date_tokens = [
        r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b",           # 12/12/2024 or 12-12-2024
        r"\b\d{1,2}\.\d{1,2}\.\d{2,4}\b",               # 12.12.2024
        r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+\d{1,2},\s+\d{4}\b",
    ]
    re_date = re.compile("|".join(date_tokens), re.I)
    re_eff = re.compile(r"effective\s*(?:as\s*of|date)?", re.I)

    max_r = min(search_rows, ws.max_row)
    max_c = min(search_cols, ws.max_column)

    def _to_mmddyyyy(s: str) -> str:
        dt = pd.to_datetime(s, errors="coerce")
        return dt.strftime("%m/%d/%Y") if pd.notna(dt) else ""

    # Pass 1: same cell contains 'effective' and the date
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v)
            if re_eff.search(s):
                m = re_date.search(s)
                if m:
                    out = _to_mmddyyyy(m.group(0))
                    if out:
                        return out
                # Often the date is in the next few cells on the same row
                for cc in range(c + 1, min(c + 5, max_c) + 1):
                    u = ws.cell(r, cc).value
                    if u is None:
                        continue
                    if hasattr(u, "year"):  # datetime/date cell
                        return u.strftime("%m/%d/%Y")
                    m2 = re_date.search(str(u))
                    if m2:
                        out = _to_mmddyyyy(m2.group(0))
                        if out:
                            return out

    # Pass 2: any early standalone date value (Excel date or string)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if hasattr(v, "year"):
                return v.strftime("%m/%d/%Y")
            m = re_date.search(str(v))
            if m:
                out = _to_mmddyyyy(m.group(0))
                if out:
                    return out

    return ""


def read_table(path: str, sheet_name: Optional[str] = None) -> Tuple[pd.DataFrame, str, str]:
    """Read the source, return (DataFrame, effective_date, sheet_title)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    effective = extract_effective_date(ws)
    header_row_1based = find_header_row(ws)
    use_header = header_row_1based - 1
    df = pd.read_excel(path, sheet_name=ws.title, header=use_header, dtype=object)
    df = df.loc[:, df.columns.notna()]
    return df, (effective or ""), ws.title


# =================================
# NPI split + Phone/Fax splitters
# =================================

def split_npi_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure 'NPI #' and 'Medicaid NPI #' exist; parse Medicaid from NPI text if embedded."""
    medicaid_col_name = "Medicaid NPI #"
    if medicaid_col_name not in df.columns:
        df[medicaid_col_name] = None

    # Find an NPI-like column (not the Medicaid one)
    npi_candidates = [c for c in df.columns if "npi" in _norm(c)]
    npi_col = None
    for c in npi_candidates:
        if "medicaid" not in _norm(c):
            npi_col = c
            break
    if npi_col is None:
        return df

    def parse_npi_pair(val):
        if pd.isna(val):
            return (None, None)
        text = str(val)
        nums = re.findall(r"\b\d{10}\b", text)
        primary = nums[0] if nums else None
        secondary = None
        if len(nums) > 1:
            secondary = nums[1]
        m = re.search(r"medicaid[^0-9]*?(\d{10})", text, re.I | re.S)
        if m:
            secondary = m.group(1)
        return primary, secondary

    need_parse = df[medicaid_col_name].isna()
    pairs = df.loc[need_parse, npi_col].apply(parse_npi_pair)
    if not pairs.empty:
        df.loc[need_parse, "NPI #"] = [p[0] for p in pairs]
        df.loc[need_parse, medicaid_col_name] = [p[1] for p in pairs]

    if npi_col != "NPI #":
        if "NPI #" not in df.columns:
            df.rename(columns={npi_col: "NPI #"}, inplace=True)
        else:
            df["NPI #"] = df["NPI #"].fillna(df[npi_col])
            df.drop(columns=[npi_col], inplace=True)

    return df


def _parse_phone_fax_text(text: object) -> Tuple[Optional[str], Optional[str]]:
    """Return (phone, fax) from combined 'P:'/'F:' text; fallback to first two numbers."""
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return (None, None)
    s = str(text)

    p = re.search(r"(?i)\bP\s*[:\-]?\s*(\(?\d{3}\)?\s*\d{3}[\s\-]?\d{4})", s)
    f = re.search(r"(?i)\bF\s*[:\-]?\s*(\(?\d{3}\)?\s*\d{3}[\s\-]?\d{4})", s)
    phone = p.group(1) if p else None
    fax = f.group(1) if f else None

    if not phone or not fax:
        nums = re.findall(r"(\(?\d{3}\)?\s*\d{3}[\s\-]?\d{4})", s)
        if not phone and len(nums) >= 1:
            phone = nums[0]
        if not fax and len(nums) >= 2:
            fax = nums[1]

    return (phone, fax)


def split_location_and_billing_phone_fax(df: pd.DataFrame) -> pd.DataFrame:
    """Split combined Phone/Fax columns into separate columns (Location and Billing)."""
    # Location combined
    loc_combined = next((c for c in df.columns if "location" in _norm(c) and "phone" in _norm(c) and "fax" in _norm(c)), None)
    if loc_combined:
        pf_series = df[loc_combined].apply(_parse_phone_fax_text)
        loc_phone = pd.Series([pf[0] for pf in pf_series], index=df.index)
        loc_fax = pd.Series([pf[1] for pf in pf_series], index=df.index)
        if "Location Phone #" not in df.columns:
            df["Location Phone #"] = None
        if "Location Fax #" not in df.columns:
            df["Location Fax #"] = None
        mask = df["Location Phone #"].isna()
        df.loc[mask, "Location Phone #"] = loc_phone[mask]
        mask = df["Location Fax #"].isna()
        df.loc[mask, "Location Fax #"] = loc_fax[mask]

    # Billing combined
    bill_combined = next((c for c in df.columns if "billing" in _norm(c) and "phone" in _norm(c) and "fax" in _norm(c)), None)
    if bill_combined:
        pf_series = df[bill_combined].apply(_parse_phone_fax_text)
        bill_phone = pd.Series([pf[0] for pf in pf_series], index=df.index)
        bill_fax = pd.Series([pf[1] for pf in pf_series], index=df.index)
        if "Billing Phone #" not in df.columns:
            df["Billing Phone #"] = None
        if "Billing Fax #" not in df.columns:
            df["Billing Fax #"] = None
        mask = df["Billing Phone #"].isna()
        df.loc[mask, "Billing Phone #"] = bill_phone[mask]
        mask = df["Billing Fax #"].isna()
        df.loc[mask, "Billing Fax #"] = bill_fax[mask]

    return df


# ==========================================
# Medicaid Provider # → dynamic columns
# ==========================================

def _split_label_number_entries(text: object) -> List[str]:
    """
    Turn a multi-line Medicaid Provider cell into:
    ['PBC 16438833', 'Medicaid NE 10026208900', 'Medicaid WY 0000000000', ...]
    Handles label and number on same line or consecutive lines.
    """
    if pd.isna(text):
        return []
    s = str(text).replace("\r", "\n")
    lines = [ln.strip() for ln in re.split(r"\n+", s) if ln and ln.strip()]

    entries: List[str] = []
    pending_label: Optional[str] = None

    for ln in lines:
        has_num = bool(re.search(r"\d{6,}", ln))

        if pending_label is None:
            if has_num and not re.fullmatch(r"[()\-\s\d]+", ln):
                entries.append(re.sub(r"\s+", " ", ln).strip())
            elif has_num:
                entries.append(ln)
            else:
                pending_label = ln
        else:
            if has_num:
                entries.append(re.sub(r"\s+", " ", f"{pending_label} {ln}").strip())
                pending_label = None
            else:
                entries.append(pending_label)
                pending_label = ln

    if pending_label:
        entries.append(pending_label)

    return [re.sub(r"\s+", " ", e).strip() for e in entries if e and e.strip()]


def expand_medicaid_provider_columns(df: pd.DataFrame,
                                     base_col: str = "Medicaid Provider #") -> pd.DataFrame:
    """Create base_col, base_col2, base_col3, ... dynamically from parsed entries."""
    if base_col not in df.columns:
        return df

    parts = df[base_col].apply(_split_label_number_entries)
    max_len = int(parts.map(len).max() or 0)
    if max_len == 0:
        return df

    out = df.copy()
    out[base_col] = parts.map(lambda x: x[0] if len(x) > 0 else None)
    for i in range(2, max_len + 1):
        out[f"{base_col}{i}"] = parts.map(lambda x, j=i-1: x[j] if len(x) > j else None)
    return out


# ==============================
# Cleanup helpers
# ==============================

def remove_trailing_notes(df: pd.DataFrame) -> pd.DataFrame:
    """Remove obvious note/footer/path rows and keep only first line of Facility."""
    df = df.copy()

    key_cols = [c for c in ["Medicaid Provider #", "Medicare Provider #", "NPI #", "Location Name"]
                if c in df.columns]
    if key_cols:
        df = df[df[key_cols].notna().any(axis=1)]

    if "Facility" in df.columns:
        note_pat = r"(?i)^\s*(note\b|[A-Z]:\\|resource\s+directory|matrix)"
        df = df[~df["Facility"].astype(str).str.strip().str.match(note_pat, na=False)]

        fac = df["Facility"].astype(str).str.replace("\r", "\n")
        fac = fac.str.split(r"\n+").str[0].str.strip()
        fac = fac.mask(fac.eq("") | fac.str.lower().eq("nan"), pd.NA)
        df["Facility"] = fac

    df = df.dropna(how="all").reset_index(drop=True)
    return df


def strip_blanks_and_drop_empty(df: pd.DataFrame) -> pd.DataFrame:
    """Convert whitespace-only strings to NA and drop all-empty rows."""
    df = df.copy()
    obj_cols = df.select_dtypes(include="object").columns
    for col in obj_cols:
        df[col] = df[col].astype(str).str.strip()
        df.loc[df[col].eq("") | df[col].str.lower().eq("nan"), col] = pd.NA
    df = df.dropna(how="all").reset_index(drop=True)
    return df


# ==============================
# Normalize, order, write
# ==============================

def normalize_and_order(df: pd.DataFrame, effective_date: str) -> pd.DataFrame:
    """Rename, split fields, expand Medicaid Provider columns, and order output."""
    rename_map: Dict[str, str] = {}
    for c in df.columns:
        n = _norm(c)
        if "facility" in n:
            rename_map[c] = "Facility"
        elif "federal" in n and "tax" in n:
            rename_map[c] = "Federal Tax ID (Pro-Fees Only)"
        elif "medicaid provider" in n:
            rename_map[c] = "Medicaid Provider #"
        elif "medicare provider" in n:
            rename_map[c] = "Medicare Provider #"
        elif n == "npi" or "npi #" in n:
            rename_map[c] = "NPI #"
        elif "location name" in n:
            rename_map[c] = "Location Name"
        elif "location address" in n:
            rename_map[c] = "Location Address"
        elif "billing name" in n and "electronic" in n:
            rename_map[c] = "Billing Name (appearing on electronic claims)"
        elif "billing name" in n and ("paper" in n or "cms 1500" in n):
            rename_map[c] = "Billing Name (appearing on paper CMS 1500)"
        elif n == "billing address":
            rename_map[c] = "Billing Address"
        elif "billing city" in n:
            rename_map[c] = "Billing City"
        elif "billing state" in n:
            rename_map[c] = "Billing State"
        elif "billing zip" in n:
            rename_map[c] = "Billing Zip"
        elif "effective date" in n:
            rename_map[c] = "Effective date"

    df = df.rename(columns=rename_map)

    # Splits
    df = split_npi_columns(df)
    df = split_location_and_billing_phone_fax(df)

    # Expand Medicaid Provider columns dynamically
    df = expand_medicaid_provider_columns(df, base_col="Medicaid Provider #")

    # Build output order (collect all Medicaid Provider #* columns)
    base = "Medicaid Provider #"
    med_cols = [base] + sorted(
        [c for c in df.columns if c.startswith(base) and c != base],
        key=lambda x: int(re.search(r"(\d+)$", x).group(1))
    )

    columns_out = (
        ["Facility", "Federal Tax ID (Pro-Fees Only)"]
        + med_cols
        + [
            "Medicare Provider #",
            "NPI #",
            "Medicaid NPI #",
            "Location Name",
            "Location Address",
            "Location Phone #",
            "Location Fax #",
            "Billing Name (appearing on electronic claims)",
            "Billing Name (appearing on paper CMS 1500)",
            "Billing Address",
            "Billing City",
            "Billing State",
            "Billing Zip",
            "Billing Phone #",
            "Billing Fax #",
            "Effective date",
        ]
    )

    # Ensure Effective date column exists/filled
    if "Effective date" not in df.columns:
        df["Effective date"] = effective_date
    else:
        df["Effective date"] = df["Effective date"].fillna(effective_date).replace("", effective_date)

    # Ensure all required columns exist
    for col in columns_out:
        if col not in df.columns:
            df[col] = pd.NA

    df_out = df[columns_out].copy()
    df_out = remove_trailing_notes(df_out)
    df_out = strip_blanks_and_drop_empty(df_out)
    return df_out


def write_output_no_banner(df_out: pd.DataFrame, out_path: str, sheet_name: str = "CO Phys - Comprehensive") -> None:
    """Write Excel starting at header row (no banner) with light header styling."""
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name=sheet_name)

        ws = writer.book[sheet_name]
        header_fill = PatternFill("solid", fgColor="DCE6F1")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx in range(1, df_out.shape[1] + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.alignment = header_alignment
            c.font = Font(bold=True)

        # Optional highlight
        highlight_fill = PatternFill("solid", fgColor="FFF2CC")
        col_index = {name: idx + 1 for idx, name in enumerate(df_out.columns)}
        for name in ["NPI #", "Medicaid NPI #", "Location Phone #", "Location Fax #"]:
            idx = col_index.get(name)
            if idx:
                ws.cell(row=1, column=idx).fill = highlight_fill

        ws.freeze_panes = "A2"

        # Widths
        widths = {
            "A": 42, "B": 24, "C": 24, "D": 24, "E": 20, "F": 16, "G": 18,
            "H": 28, "I": 40, "J": 18, "K": 18, "L": 36, "M": 36, "N": 24,
            "O": 16, "P": 10, "Q": 12, "R": 18, "S": 18, "T": 16, "U": 16, "V": 16
        }
        # guard in case there are many dynamic columns
        for i in range(1, ws.max_column + 1):
            col_letter = chr(ord('A') + i - 1)
            ws.column_dimensions[col_letter].width = widths.get(col_letter, 18)


# ==============================
# Pipeline
# ==============================

def transform(src_path: str, dst_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    df_src, eff, title = read_table(src_path, sheet_name)
    df_out = normalize_and_order(df_src, eff)
    write_output_no_banner(df_out, dst_path, sheet_name=title or "CO Phys - Comprehensive")
    return df_out


# ============
# Run locally
# ============

if __name__ == "__main__":
    # TODO: update these paths for your machine
    SRC = r"H:\Transformation\2025 Aug Week4\Data\CO_NonDLG_Banner_Health_...source.xlsx"
    DST = r"H:\Transformation\2025 Aug Week4\Data\CO_NonDLG_Banner_Health_...output.xlsx"

    df_result = transform(SRC, DST)
    print(f"Saved: {DST}")
    # quick preview in notebooks:
    try:
        from IPython.display import display
        display(df_result.head(10))
    except Exception:
        print(df_result.head(10).to_string(index=False))
