# Fix fillna(list) issue by masking assignments
import pandas as pd
import re
from typing import Optional, Tuple, Dict, List
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill

def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s).strip().lower())

def find_header_row(ws: Worksheet, search_terms: List[str] = None, scan_rows: int = 50) -> int:
    if search_terms is None:
        search_terms = ["facility","npi","medicare provider","location name","billing name"]
    norm_terms = set(_norm(t) for t in search_terms)
    best_row = None
    best_hits = -1
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        vals = [c.value for c in ws[r]]
        norms = [_norm(v) for v in vals if v is not None]
        hits = sum(any(t in n for t in norm_terms) for n in norms)
        if hits > best_hits:
            best_hits = hits; best_row = r
    return best_row or 6

def extract_effective_date(ws: Worksheet) -> Optional[str]:
    date_pat = re.compile(r"effective\s+as\s+of\s*[:\-]?\s*(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})", re.I)
    for r in range(1, min(12, ws.max_row) + 1):
        row_text = " ".join(str(c.value) for c in ws[r] if c.value is not None)
        m = date_pat.search(row_text)
        if m: return m.group(1)
    return None

def read_table(path: str, sheet_name: Optional[str] = None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    effective = extract_effective_date(ws)
    header_row_1based = find_header_row(ws)
    use_header = header_row_1based - 1
    df = pd.read_excel(path, sheet_name=ws.title, header=use_header, dtype=object)
    df = df.loc[:, df.columns.notna()]
    return df, (effective or "")

def split_npi_columns(df: pd.DataFrame) -> pd.DataFrame:
    medicaid_col_name = "Medicaid NPI #"
    if medicaid_col_name not in df.columns:
        df[medicaid_col_name] = None
    npi_candidates = [c for c in df.columns if "npi" in _norm(c)]
    npi_col = None
    for c in npi_candidates:
        if "medicaid" not in _norm(c):
            npi_col = c; break
    if npi_col is None: return df
    def parse_npi_pair(val):
        if pd.isna(val): return (None, None)
        text = str(val)
        nums = re.findall(r"\b\d{10}\b", text)
        primary = nums[0] if nums else None
        secondary = None
        if len(nums) > 1: secondary = nums[1]
        m = re.search(r"medicaid[^0-9]*?(\d{10})", text, re.I | re.S)
        if m: secondary = m.group(1)
        return primary, secondary
    need_parse = df[medicaid_col_name].isna()
    pairs = df.loc[need_parse, npi_col].apply(parse_npi_pair)
    if not pairs.empty:
        df.loc[need_parse, "NPI #"] = [p[0] for p in pairs]
        df.loc[need_parse, medicaid_col_name] = [p[1] for p in pairs]
    if npi_col != "NPI #":
        if "NPI #" not in df.columns:
            df.rename(columns={npi_col: "NPI #"}, inplace=True)
        else:
            df["NPI #"] = df["NPI #"].fillna(df[npi_col]); df.drop(columns=[npi_col], inplace=True)
    return df

def _parse_phone_fax_text(text: str) -> Tuple[Optional[str], Optional[str]]:
    if text is None or (isinstance(text, float) and pd.isna(text)): return (None, None)
    s = str(text)
    p = re.search(r"(?i)\bP\s*[:\-]?\s*(\(?\d{3}\)?\s*\d{3}[\s\-]?\d{4})", s)
    f = re.search(r"(?i)\bF\s*[:\-]?\s*(\(?\d{3}\)?\s*\d{3}[\s\-]?\d{4})", s)
    phone = p.group(1) if p else None
    fax = f.group(1) if f else None
    if not phone or not fax:
        nums = re.findall(r"(\(?\d{3}\)?\s*\d{3}[\s\-]?\d{4})", s)
        if not phone and len(nums) >= 1: phone = nums[0]
        if not fax and len(nums) >= 2: fax = nums[1]
    return (phone, fax)

def split_location_and_billing_phone_fax(df: pd.DataFrame) -> pd.DataFrame:
    # LOCATION combined
    loc_combined = None
    for c in df.columns:
        n = _norm(c)
        if "location" in n and "phone" in n and "fax" in n:
            loc_combined = c; break
    if loc_combined:
        pf_series = df[loc_combined].apply(_parse_phone_fax_text)
        loc_phone = pd.Series([pf[0] for pf in pf_series], index=df.index)
        loc_fax = pd.Series([pf[1] for pf in pf_series], index=df.index)
        if "Location Phone #" not in df.columns: df["Location Phone #"] = None
        if "Location Fax #" not in df.columns: df["Location Fax #"] = None
        mask = df["Location Phone #"].isna(); df.loc[mask, "Location Phone #"] = loc_phone[mask]
        mask = df["Location Fax #"].isna(); df.loc[mask, "Location Fax #"] = loc_fax[mask]

    # BILLING combined
    bill_combined = None
    for c in df.columns:
        n = _norm(c)
        if "billing" in n and "phone" in n and "fax" in n:
            bill_combined = c; break
    if bill_combined:
        pf_series = df[bill_combined].apply(_parse_phone_fax_text)
        bill_phone = pd.Series([pf[0] for pf in pf_series], index=df.index)
        bill_fax = pd.Series([pf[1] for pf in pf_series], index=df.index)
        if "Billing Phone #" not in df.columns: df["Billing Phone #"] = None
        if "Billing Fax #" not in df.columns: df["Billing Fax #"] = None
        mask = df["Billing Phone #"].isna(); df.loc[mask, "Billing Phone #"] = bill_phone[mask]
        mask = df["Billing Fax #"].isna(); df.loc[mask, "Billing Fax #"] = bill_fax[mask]

    return df

def normalize_and_order(df: pd.DataFrame, effective_date: str) -> pd.DataFrame:
    rename_map: Dict[str, str] = {}
    for c in df.columns:
        n = _norm(c)
        if "facility" in n: rename_map[c] = "Facility"
        elif "federal" in n and "tax" in n: rename_map[c] = "Federal Tax ID (Pro-Fees Only)"
        elif "medicaid provider" in n: rename_map[c] = "Medicaid Provider #"
        elif "medicare provider" in n: rename_map[c] = "Medicare Provider #"
        elif n == "npi" or "npi #" in n: rename_map[c] = "NPI #"
        elif "location name" in n: rename_map[c] = "Location Name"
        elif "location address" in n: rename_map[c] = "Location Address"
        elif "billing name" in n and "electronic" in n: rename_map[c] = "Billing Name (appearing on electronic claims)"
        elif "billing name" in n and ("paper" in n or "cms 1500" in n): rename_map[c] = "Billing Name (appearing on paper CMS 1500)"
        elif n == "billing address": rename_map[c] = "Billing Address"
        elif "billing city" in n: rename_map[c] = "Billing City"
        elif "billing state" in n: rename_map[c] = "Billing State"
        elif "billing zip" in n: rename_map[c] = "Billing Zip"
        elif "effective date" in n: rename_map[c] = "Effective date"
    df = df.rename(columns=rename_map)
    df = split_npi_columns(df)
    df = split_location_and_billing_phone_fax(df)
    columns_out = [
        "Facility","Federal Tax ID (Pro-Fees Only)","Medicaid Provider #","Medicare Provider #",
        "NPI #","Medicaid NPI #","Location Name","Location Address",
        "Location Phone #","Location Fax #",
        "Billing Name (appearing on electronic claims)","Billing Name (appearing on paper CMS 1500)",
        "Billing Address","Billing City","Billing State","Billing Zip",
        "Billing Phone #","Billing Fax #","Effective date"
    ]
    if "Effective date" not in df.columns: df["Effective date"] = effective_date
    else: df["Effective date"] = df["Effective date"].fillna(effective_date).replace("", effective_date)
    for col in columns_out:
        if col not in df.columns: df[col] = None
    return df[columns_out].copy()

def write_output_no_banner(df_out: pd.DataFrame, path_out: str):
    with pd.ExcelWriter(path_out, engine="openpyxl") as writer:
        sheet = "CO Phys - Comprehensive"
        df_out.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.book[sheet]
        header_fill = PatternFill("solid", fgColor="DCE6F1")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, df_out.shape[1] + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill; c.alignment = header_alignment; c.font = Font(bold=True)
        highlight_fill = PatternFill("solid", fgColor="FFF2CC")
        col_index = {name: idx + 1 for idx, name in enumerate(df_out.columns)}
        for name in ["NPI #","Medicaid NPI #","Location Phone #","Location Fax #"]:
            idx = col_index.get(name)
            if idx:
                ws.cell(row=1, column=idx).fill = highlight_fill
        ws.freeze_panes = "A2"
        widths = {"A":42,"B":24,"C":20,"D":20,"E":16,"F":18,"G":28,"H":40,"I":18,"J":18,"K":36,"L":36,"M":24,"N":16,"O":10,"P":12,"Q":18,"R":18,"S":16}
        for col_letter, width in widths.items():
            if ws.max_column >= ord(col_letter)-ord("A")+1:
                ws.column_dimensions[col_letter].width = width

def transform(src_path: str, out_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    df_src, eff = read_table(src_path, sheet_name)
    df_out = normalize_and_order(df_src, eff)
    write_output_no_banner(df_out, out_path)
    return df_out

src = r"C:\Users\Nitesh\Exl_Work\2025\Aug\week1\Data\Colorado_Billing_Information_Matrix_v2_source.xlsx"
dst = r"C:\Users\Nitesh\Exl_Work\2025\Aug\week1\Data\Colorado_Billing_Information_Matrix_v2_output1.xlsx"
df_result = transform(src, dst)
dst
